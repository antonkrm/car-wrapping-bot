from aiogram import Router, F, types, Bot
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.types.input_file import BufferedInputFile
from datetime import datetime, timedelta
import re
import io
from io import BytesIO
from aiogram.types import InputFile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import zipfile
import tempfile
import os
import asyncio
import aiohttp

from db import (
    add_user, set_admin, is_admin, get_user_id,
    add_report, add_car, add_photo, get_photos_by_date, get_connection, get_photos_by_month
)
from parser import parse_report_text
from config import ADMIN_PASSWORD

router = Router()
photo_buffer = {}  # временное хранилище фото: {user_id: [file_id, ...]}


class AdminLogin(StatesGroup):
    password = State()


class Form(StatesGroup):
    waiting_for_report_date = State()
    waiting_for_photo_date = State()


admin_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Отчет"), KeyboardButton(text="📸 Фото")],
    ],
    resize_keyboard=True
)


@router.message(CommandStart())
async def cmd_start(message: types.Message):
    add_user(message.from_user.id, message.from_user.full_name)
    if is_admin(message.from_user.id):
        await message.answer("Привет, админ! Отправь отчет или выбери действие:", reply_markup=admin_keyboard)
    else:
        await message.answer("Привет! Отправь отчет в свободной форме. Фото тоже можешь прислать отдельным сообщением.",
                             reply_markup=ReplyKeyboardRemove())


@router.message(Command("admin"))
async def admin_login(message: types.Message, state: FSMContext):
    await state.set_state(AdminLogin.password)
    await message.answer("Введите пароль администратора:")


@router.message(AdminLogin.password)
async def check_password(message: types.Message, state: FSMContext):
    if message.text == ADMIN_PASSWORD:
        set_admin(message.from_user.id)
        await message.answer("✅ Администратор подтвержден.", reply_markup=admin_keyboard)
    else:
        await message.answer("❌ Неверный пароль.")
    await state.clear()


@router.message(F.photo)
async def handle_photo(message: types.Message):
    user_id = message.from_user.id
    file_id = message.photo[-1].file_id

    if message.caption:
        try:
            cars, date = parse_report_text(message.caption)
            uid = get_user_id(user_id)
            report_id = add_report(uid, date)

            for car in cars:
                add_car(report_id, car["plate"], car["description"], car["area"], car["cost"], car["labor_cost"])

            add_photo(report_id, file_id)

            total_labor = sum(c['labor_cost'] for c in cars)
            await message.answer(
                f"✅ Отчет с фото за {date} принят. Машин: {len(cars)}\n"
                f"🔧 Общая стоимость работ: {total_labor} ₽",
                reply_markup=admin_keyboard if is_admin(user_id) else ReplyKeyboardRemove()
            )
        except Exception as e:
            await message.answer(f"⚠️ Ошибка при обработке отчета с фото: {e}")
    else:
        photo_buffer.setdefault(user_id, []).append(file_id)
        await message.answer("📸 Фото получено. Пришли теперь текст отчета, чтобы связать с фото.")


@router.message(F.text == "📸 Фото")
async def photo_by_date_request(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return await message.answer("❌ Команда только для администратора.")
    await state.set_state(Form.waiting_for_photo_date)
    await message.answer("Введите дату или месяц для выгрузки фото (формат: ГГГГ-ММ-ДД или ГГГГ-ММ):")


@router.message(Form.waiting_for_photo_date)
async def handle_photo_by_date(message: types.Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        await message.answer("❌ Команда только для администратора.")
        await state.clear()
        return

    date_input = message.text.strip()
    
    # Проверка формата даты (ГГГГ-ММ-ДД)
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_input):
        photos = get_photos_by_date(date_input)
        if not photos:
            await message.answer("Нет фото за эту дату.")
        else:
            media = [types.InputMediaPhoto(media=pid) for pid in photos[:10]]
            if len(media) == 1:
                await message.answer_photo(media[0].media)
            else:
                await message.answer_media_group(media)
    # Проверка формата месяца (ГГГГ-ММ)
    elif re.match(r"^\d{4}-\d{2}$", date_input):
        year, month = map(int, date_input.split("-"))
        photos = get_photos_by_month(year, month)
        if not photos:
            await message.answer("Нет фото за этот месяц.")
        else:
            # Создание архива фотографий
            zip_buffer = io.BytesIO()
            
            async with aiohttp.ClientSession() as session:
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for i, file_id in enumerate(photos, 1):
                        try:
                            # Получение информации о файле
                            file = await bot.get_file(file_id)
                            file_path = file.file_path
                            
                            # Загрузка файла
                            file_url = f"https://api.telegram.org/file/bot{bot.token}/{file_path}"
                            async with session.get(file_url) as resp:
                                if resp.status == 200:
                                    file_content = await resp.read()
                                    # Определяем расширение файла по file_path
                                    ext = file_path.split('.')[-1] if '.' in file_path else 'jpg'
                                    zip_file.writestr(f"photo_{i:03d}.{ext}", file_content)
                                else:
                                    # Если не удалось загрузить, добавляем текстовый файл с информацией
                                    zip_file.writestr(f"photo_{i:03d}_{file_id}.txt", f"Фото с ID: {file_id}\nНе удалось загрузить файл. Статус: {resp.status}")
                        except Exception as e:
                            # Если произошла ошибка, добавляем текстовый файл с информацией
                            zip_file.writestr(f"photo_{i:03d}_{file_id}.txt", f"Фото с ID: {file_id}\nОшибка при загрузке: {str(e)}")
            
            zip_buffer.seek(0)
            input_file = BufferedInputFile(zip_buffer.read(), filename=f"photos_{year}_{month:02d}.zip")
            await message.answer_document(input_file, caption=f"Архив фотографий за {date_input} ({len(photos)} шт.)")
    else:
        await message.answer("❌ Неверный формат даты. Введите в формате ГГГГ-ММ-ДД или ГГГГ-ММ.")
    
    await state.clear()


@router.message(F.text == "📊 Отчет")
async def report_by_date_request(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return await message.answer("❌ Команда только для администратора.")
    await state.set_state(Form.waiting_for_report_date)
    await message.answer("Введите дату или месяц для отчета (формат: ГГГГ-ММ или ГГГГ-ММ-ДД):")


@router.message(Form.waiting_for_report_date)
async def handle_report_by_date(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        await message.answer("❌ Команда только для администратора.")
        await state.clear()
        return

    date_arg = message.text.strip()

    if re.match(r"^\d{4}-\d{2}-\d{2}$", date_arg):
        date_from = date_to = date_arg
    elif re.match(r"^\d{4}-\d{2}$", date_arg):
        year, month = map(int, date_arg.split("-"))
        date_from = f"{year}-{month:02d}-01"
        if month == 12:
            date_to = f"{year + 1}-01-01"
        else:
            date_to = f"{year}-{month + 1:02d}-01"
    else:
        await message.answer("❌ Неверный формат даты.")
        return

    if date_to.endswith("-01"):
        date_to_dt = datetime.strptime(date_to, "%Y-%m-%d") - timedelta(days=1)
        date_to = date_to_dt.strftime("%Y-%m-%d")

    count, area, cost, labor_cost = get_report_summary(date_from, date_to)
    cars = get_car_details(date_from, date_to)

    msg = (
        f"📊 Отчет с {date_from} по {date_to}\n"
        f"🚗 Машин оклеено: {count or 0}\n"
        f"📏 Площадь пленки (м²): {round(area or 0, 2)}\n"
        f"💰 Стоимость материалов (руб): {int(cost or 0)}\n"
        f"🔧 Общая стоимость работ (руб): {int(labor_cost or 0)}\n"
    )

    details = generate_cars_report(cars)
    full_msg = msg + details
    if len(full_msg) < 3900:
        await message.answer(full_msg)
    else:
        await message.answer(msg)
        await message.answer(details[:3900])

    if cars:
        excel_stream = create_excel_report(cars, f"{date_from} — {date_to}")
        excel_stream.seek(0)
        await message.answer(f"Файл подготовлен, размер: {len(excel_stream.getvalue())} байт")  # отладка

        file_bytes = excel_stream.read()
        input_file = BufferedInputFile(file_bytes, filename=f"report_{date_from}_to_{date_to}.xlsx")
        await message.answer_document(input_file, caption="📊 Ваш Excel-отчет по машинам")

    await state.clear()


@router.message(F.text)
async def handle_text_report(message: types.Message):
    if message.text.startswith("/"):
        return
    user_id = message.from_user.id
    add_user(user_id, message.from_user.full_name)

    try:
        cars, date = parse_report_text(message.text)
        uid = get_user_id(user_id)
        report_id = add_report(uid, date)

        for car in cars:
            add_car(report_id, car["plate"], car["description"], car["area"], car["cost"], car["labor_cost"])

        if user_id in photo_buffer:
            for file_id in photo_buffer[user_id]:
                add_photo(report_id, file_id)
            del photo_buffer[user_id]

        total_labor = sum(c['labor_cost'] for c in cars)
        await message.answer(
            f"✅ Отчет за {date} принят. Машин: {len(cars)}\n"
            f"🔧 Общая стоимость работ: {total_labor} ₽",
            reply_markup=admin_keyboard if is_admin(user_id) else ReplyKeyboardRemove()
        )
    except Exception as e:
        await message.answer(f"⚠️ Ошибка при обработке отчета: {e}")


# --- Вспомогательные функции для отчетов и Excel ---

def get_report_summary(date_from: str, date_to: str, user_id=None):
    with get_connection() as conn:
        cur = conn.cursor()
        query = """
            SELECT COUNT(DISTINCT license_plate), 
                   SUM(area), 
                   SUM(cost),
                   SUM(labor_cost)
            FROM cars
            JOIN reports ON cars.report_id = reports.id
            JOIN users ON reports.user_id = users.id
            WHERE date(reports.date) BETWEEN ? AND ?
        """
        params = [date_from, date_to]
        if user_id:
            query += " AND users.tg_id = ?"
            params.append(user_id)
        cur.execute(query, params)
        row = cur.fetchone()
        return row or (0, 0, 0, 0)


def get_car_details(date_from: str, date_to: str, user_id=None):
    with get_connection() as conn:
        cur = conn.cursor()
        query = """
            SELECT cars.license_plate, cars.description, cars.area, cars.cost, cars.labor_cost, reports.date, users.name
            FROM cars
            JOIN reports ON cars.report_id = reports.id
            JOIN users ON reports.user_id = users.id
            WHERE date(reports.date) BETWEEN ? AND ?
        """
        params = [date_from, date_to]
        if user_id:
            query += " AND users.tg_id = ?"
            params.append(user_id)
        cur.execute(query, params)
        rows = cur.fetchall()
        cols = [desc[0] for desc in cur.description]
        return [dict(zip(cols, row)) for row in rows]


def generate_cars_report(cars: list) -> str:
    if not cars:
        return "Нет данных по машинам за выбранный период."
    lines = ["\nДетализация:"]
    for idx, car in enumerate(cars, 1):
        lines.append(
            f"{idx}. {car['license_plate']} — {car['description']}\n"
            f"    Площадь: {car['area']:.2f} м² | Материалы: {int(car['cost'])} ₽ | Работы: {int(car['labor_cost'])} ₽"
            + (f"\n    Исполнитель: {car['name']}" if 'name' in car and car['name'] else "")
            + (f"\n    Дата: {car['date']}" if 'date' in car and car['date'] else "")
        )
    return '\n'.join(lines)


def create_excel_report(cars: list, date_range: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"

    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    border_side = Side(style="thin", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers = ["№", "Номер", "Описание", "Площадь (м²)", "Материалы (руб)", "Работы (руб)", "Дата", "Исполнитель"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, car in enumerate(cars, 1):
        row = [
            idx,
            car.get("license_plate", ""),
            car.get("description", ""),
            round(car.get("area", 0), 2),
            int(car.get("cost", 0)),
            int(car.get("labor_cost", 0)),
            car.get("date", ""),
            car.get("name", ""),
        ]
        ws.append(row)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
